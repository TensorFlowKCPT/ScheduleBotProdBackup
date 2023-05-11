from requests import *
from bs4 import *
import os
import pandas
import openpyxl

def Download():
    url = 'https://kcpt72.ru/schedule/'
    urldownload = ParsKCPT(url)

    excelfile = []
    wordfile = ""
    for i in urldownload:
        filename = str(i).split('/')
        if filename[3] == 'document':
            wordfile = f"https://docs.google.com/document/u/0/export?format=docx&id={filename[5]}&token=AC4w5VhQRMQWeInMWM3h4AvaXLoxHWZsPw%3A1683047241502&includes_info_params=true&usp=sharing&cros_files=false&inspectorResult=%7B%22pc%22%3A9%2C%22lplc%22%3A10%7D"
        if filename[3] == 'spreadsheets':
            excelfile.append(f"https://drive.google.com/u/0/uc?id={filename[5]}&export=download")

    path = "ExcelAndWord"

    num = 1
    for i in excelfile:
        response = get(i)
        # сохраните файл в папке с заданным именем
        with open(os.path.join(path, f'Schedule{num}.xlsx'), 'wb') as f:
            f.write(response.content)
        num += 1

    response = get(wordfile)
    # сохраните файл в папке с заданным именем
    with open(os.path.join(path, 'Change.docx'), 'wb') as f:
        f.write(response.content)
    
    twotable("ExcelAndWord/Schedule2.xlsx", "ExcelAndWord/Schedule1.xlsx")

def ParsKCPT(url):

    arr_a_google = []
    response = get(url)
    html_content = response.text


    soup = BeautifulSoup(html_content, 'html.parser')
    a_tags = soup.find_all("a")

    for i in a_tags:
        if "docs.google.com" in str(i):
            ps = BeautifulSoup(str(i), "html.parser")
            link = ps.find('a')['href']
            arr_a_google.append(link)
    
    return arr_a_google

def twotable(pathfile1: str, pathfile2: str):
    # Читаем первый файл
    workbook1 = openpyxl.load_workbook(pathfile1, data_only=True)

    # Читаем второй файл
    workbook2 = openpyxl.load_workbook(pathfile2, data_only=True)

    # Создаем новый Workbook
    merged_workbook = openpyxl.Workbook()
    merged_sheet = merged_workbook.active

    # Сливаем данные из первого файла
    for sheet_name in workbook1.sheetnames:
        sheet = workbook1[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            merged_sheet.append(row)

    # Сливаем данные из второго файла
    for sheet_name in workbook2.sheetnames:
        sheet = workbook2[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            merged_sheet.append(row)

    # Сохраняем объединенный Workbook в новый файл
    merged_workbook.save("ExcelAndWord/2table.xlsx")

Download()









